"""
Vistas para gestión de notas y exámenes por profesores
"""
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from datetime import datetime, date # Se incluye date
from django.utils.dateparse import parse_date # Se incluye para parsear fechas POST
from django.core.exceptions import ValidationError # Se incluye para manejar errores de modelo
from decimal import Decimal
import json

# Imports de Modelos
from app.models.usuario.models import Profesor, Usuario, Estudiante
from app.models.curso.models import Curso
from app.models.evaluacion.models import Nota, FechaExamen # Se incluye FechaExamen
from app.models.matricula.models import Matricula
from app.models.horario.models import Horario
from services.notasService import NotasService


notasService = NotasService()


def decimal_default(obj):
    """Convertir Decimal a float para JSON serialization"""
    if isinstance(obj, Decimal):
        return float(obj)
    raise TypeError


@login_required
def seleccionar_curso_notas(request):
    """
    Vista para que el profesor seleccione el curso para ingresar notas
    Permite ingresar notas a cualquier profesor asignado al curso
    """
    try:
        profesor = Profesor.objects.get(usuario=request.user)
        
        # Obtener todos los cursos donde el profesor está asignado (cualquier tipo de clase)
        horarios = Horario.objects.filter(
            profesor=profesor,
            is_active=True,
            periodo_academico='2025-B'
        ).select_related('curso')
        
        # Eliminar duplicados manualmente (compatible con SQLite)
        cursos_dict = {}
        for h in horarios:
            if h.curso.codigo not in cursos_dict:
                cursos_dict[h.curso.codigo] = h.curso
        
        cursos = list(cursos_dict.values())
        
        if not cursos:
            messages.warning(request, 'No tiene cursos asignados. Debe estar asignado a un curso para poder ingresar notas.')
        
        context = {
            'profesor': profesor,
            'cursos': cursos
        }
        
        return render(request, 'profesor/seleccionar_curso_notas.html', context)
        
    except Profesor.DoesNotExist:
        messages.error(request, 'Solo los profesores pueden acceder a esta página.')
        return redirect('login')


@login_required
def ingresar_notas(request, curso_codigo, unidad):
    """
    Vista para ingresar notas de una unidad específica
    
    Reglas:
    - Profesor de TEORÍA: puede subir PARCIAL y CONTINUA
    - Profesor de PRÁCTICA: 
        * Si curso NO tiene teoría → puede subir PARCIAL y CONTINUA
        * Si curso SÍ tiene teoría → solo puede subir CONTINUA
    - Profesor de LABORATORIO: NO puede subir notas
    """
    try:
        profesor = Profesor.objects.get(usuario=request.user)
        curso = get_object_or_404(Curso, codigo=curso_codigo)
        
        # Verificar qué tipo de clase dicta este profesor en este curso
        horarios_profesor = Horario.objects.filter(
            profesor=profesor,
            curso=curso,
            is_active=True,
            periodo_academico='2025-B'
        )
        
        if not horarios_profesor.exists():
            messages.error(request, 'No está asignado a este curso.')
            return redirect('seleccionar_curso_notas')
        
        # Determinar el tipo de profesor
        es_profesor_teoria = horarios_profesor.filter(tipo_clase='TEORIA').exists()
        es_profesor_practica = horarios_profesor.filter(tipo_clase='PRACTICA').exists()
        es_profesor_laboratorio = horarios_profesor.filter(tipo_clase='LABORATORIO').exists()
        
        # Laboratorio no puede subir notas
        if es_profesor_laboratorio and not es_profesor_teoria and not es_profesor_practica:
            messages.error(request, 'Los profesores de laboratorio no pueden ingresar notas. Debe coordinar con el profesor de teoría o práctica.')
            return redirect('seleccionar_curso_notas')
        
        # Verificar si el curso tiene teoría
        curso_tiene_teoria = Horario.objects.filter(
            curso=curso,
            tipo_clase='TEORIA',
            is_active=True,
            periodo_academico='2025-B'
        ).exists()
        
        # Determinar permisos según las reglas
        puede_subir_parcial = False
        puede_subir_continua = False
        
        if es_profesor_teoria:
            # Profesor de teoría: puede subir todo
            puede_subir_parcial = True
            puede_subir_continua = True
        elif es_profesor_practica:
            if curso_tiene_teoria:
                # Curso con teoría: profesor de práctica solo sube continua
                puede_subir_parcial = False
                puede_subir_continua = True
            else:
                # Curso sin teoría: profesor de práctica sube todo
                puede_subir_parcial = True
                puede_subir_continua = True
        
        # Verificar permiso según el tipo de evaluación de la unidad
        if unidad in [1, 3]:  # Parciales
            if not puede_subir_parcial:
                messages.error(request, 'Solo el profesor de teoría puede ingresar notas de parcial. Los profesores de práctica solo pueden ingresar notas de continua.')
                return redirect('seleccionar_curso_notas')
        elif unidad in [2, 4]:  # Continuas
            if not puede_subir_continua:
                messages.error(request, 'No tiene permisos para ingresar notas de continua.')
                return redirect('seleccionar_curso_notas')
        
        if request.method == 'POST':
            # Procesar notas enviadas
            notas_data = []
            
            # Obtener estudiantes del formulario
            estudiantes_codigos = request.POST.getlist('estudiante_codigo[]')
            
            for est_codigo in estudiantes_codigos:
                nota_parcial = request.POST.get(f'nota_parcial_{est_codigo}')
                nota_continua = request.POST.get(f'nota_continua_{est_codigo}')
                archivo_examen = request.FILES.get(f'archivo_examen_{est_codigo}')
                
                if nota_parcial or nota_continua:
                    notas_data.append({
                        'estudiante_codigo': est_codigo,
                        'nota_parcial': float(nota_parcial) if nota_parcial else None,
                        'nota_continua': float(nota_continua) if nota_continua else None,
                        'archivo_examen': archivo_examen
                    })
            
            # Ingresar notas usando el servicio
            resultado = notasService.ingresarNotas(
                curso_codigo=curso_codigo,
                profesor_usuario_id=profesor.usuario.codigo,
                unidad=unidad,
                notas_data=notas_data
            )
            
            if resultado['success']:
                messages.success(
                    request, 
                    f'Notas registradas: {resultado["notas_creadas"]} creadas, {resultado["notas_actualizadas"]} actualizadas.'
                )
                
                if resultado['errores']:
                    for error in resultado['errores']:
                        messages.warning(request, error)
                
                # Establecer flag para mostrar pop-up de recordatorio
                request.session['mostrar_recordatorio_notas'] = True
                
                # Redirigir a la misma página para mostrar el modal
                return redirect('ingresar_notas', curso_codigo=curso_codigo, unidad=unidad)
            else:
                messages.error(request, 'Error al registrar notas.')
        
        # GET: Mostrar formulario
        matriculas = notasService.obtenerEstudiantesParaNotas(
            curso_codigo=curso_codigo,
            profesor_usuario_id=profesor.usuario.codigo
        )
        
        # Obtener notas existentes para esta unidad
        notas_existentes = {}
        for matricula in matriculas:
            estudiante = matricula.estudiante
            
            # Buscar nota parcial existente
            nota_parcial = Nota.objects.filter(
                curso=curso,
                estudiante=estudiante,
                categoria='PARCIAL',
                unidad=unidad,
                numero_evaluacion=1
            ).first()
            
            # Buscar nota continua existente
            nota_continua = Nota.objects.filter(
                curso=curso,
                estudiante=estudiante,
                categoria='CONTINUA',
                unidad=unidad,
                numero_evaluacion=1
            ).first()
            
            # Usar el usuario.codigo como clave (que coincide con el código usado en el formulario)
            notas_existentes[estudiante.usuario.codigo] = {
                'parcial': nota_parcial,
                'continua': nota_continua
            }
        
        context = {
            'curso': curso,
            'profesor': profesor,
            'unidad': unidad,
            'matriculas': matriculas,
            'notas_existentes': notas_existentes,
            'fecha_actual': timezone.now(),
            'mostrar_recordatorio': request.session.pop('mostrar_recordatorio_notas', False),
            'puede_subir_parcial': puede_subir_parcial,
            'puede_subir_continua': puede_subir_continua,
            'es_profesor_teoria': es_profesor_teoria,
            'es_profesor_practica': es_profesor_practica
        }
        
        return render(request, 'profesor/ingresar_notas.html', context)
        
    except Profesor.DoesNotExist:
        messages.error(request, 'Solo los profesores pueden acceder a esta página.')
        return redirect('login')
    except Exception as e:
        messages.error(request, f'Error: {str(e)}')
        return redirect('seleccionar_curso_notas')


@login_required
def estadisticas_notas(request, curso_codigo):
    """
    Vista para ver estadísticas de notas de un curso
    Profesores de teoría o práctica pueden ver estadísticas
    """
    try:
        profesor = Profesor.objects.get(usuario=request.user)
        curso = get_object_or_404(Curso, codigo=curso_codigo)
        
        # Verificar que el profesor esté asignado al curso (teoría o práctica)
        horarios_profesor = Horario.objects.filter(
            profesor=profesor,
            curso=curso,
            is_active=True,
            periodo_academico='2025-B'
        )
        
        if not horarios_profesor.exists():
            messages.error(request, 'No está asignado a este curso.')
            return redirect('seleccionar_curso_notas')
        
        # Verificar que no sea solo laboratorio
        es_solo_laboratorio = (
            horarios_profesor.filter(tipo_clase='LABORATORIO').exists() and
            not horarios_profesor.filter(tipo_clase__in=['TEORIA', 'PRACTICA']).exists()
        )
        
        if es_solo_laboratorio:
            messages.error(request, 'Los profesores de laboratorio no tienen acceso a estadísticas. Debe coordinar con el profesor de teoría o práctica.')
            return redirect('seleccionar_curso_notas')
        
        unidad = int(request.GET.get('unidad', 1))
        
        # Obtener estadísticas
        stats_parcial = notasService.calcularEstadisticas(curso_codigo, unidad, 'PARCIAL')
        stats_continua = notasService.calcularEstadisticas(curso_codigo, unidad, 'CONTINUA')
        
        # Datos para gráficas
        datos_grafica = notasService.obtenerNotasParaGrafica(curso_codigo, unidad)
        
        context = {
            'profesor': profesor,
            'curso': curso,
            'unidad': unidad,
            'stats_parcial': stats_parcial,
            'stats_continua': stats_continua,
            'datos_grafica': json.dumps(datos_grafica, default=decimal_default),
            'unidades': [1, 2, 3]
        }
        
        return render(request, 'profesor/estadisticas_notas.html', context)
        
    except Profesor.DoesNotExist:
        messages.error(request, 'Solo los profesores pueden acceder a esta página.')
        return redirect('login')
    except Exception as e:
        messages.error(request, f'Error: {str(e)}')
        return redirect('profesor_dashboard')


@login_required
def generar_reporte_secretaria(request, curso_codigo, unidad):
    """
    Vista para generar reporte de notas para secretaría
    Profesores de teoría o práctica pueden generar reportes
    Permite subir archivos de exámenes
    """
    try:
        profesor = Profesor.objects.get(usuario=request.user)
        curso = get_object_or_404(Curso, codigo=curso_codigo)
        
        # Verificar que el profesor esté asignado al curso (teoría o práctica)
        horarios_profesor = Horario.objects.filter(
            profesor=profesor,
            curso=curso,
            is_active=True,
            periodo_academico='2025-B'
        )
        
        if not horarios_profesor.exists():
            messages.error(request, 'No está asignado a este curso.')
            return redirect('seleccionar_curso_notas')
        
        # Verificar que no sea solo laboratorio
        es_solo_laboratorio = (
            horarios_profesor.filter(tipo_clase='LABORATORIO').exists() and
            not horarios_profesor.filter(tipo_clase__in=['TEORIA', 'PRACTICA']).exists()
        )
        
        if es_solo_laboratorio:
            messages.error(request, 'Los profesores de laboratorio no pueden generar reportes. Debe coordinar con el profesor de teoría o práctica.')
            return redirect('seleccionar_curso_notas')
        
        # Manejar POST para subida de archivos
        if request.method == 'POST':
            estudiantes_codigos = request.POST.getlist('estudiante_codigo[]')
            archivos_subidos = 0
            
            for est_codigo in estudiantes_codigos:
                archivo_examen = request.FILES.get(f'archivo_examen_{est_codigo}')
                
                if archivo_examen:
                    try:
                        # Buscar la nota parcial del estudiante
                        from app.models.usuario.models import Usuario
                        estudiante_usuario = Usuario.objects.get(codigo=est_codigo)
                        estudiante = estudiante_usuario.estudiante
                        
                        nota_parcial = Nota.objects.filter(
                            curso=curso,
                            estudiante=estudiante,
                            categoria='PARCIAL',
                            unidad=unidad
                        ).first()
                        
                        if nota_parcial:
                            nota_parcial.archivo_examen = archivo_examen
                            nota_parcial.save()
                            archivos_subidos += 1
                    except Exception as e:
                        messages.warning(request, f'Error al subir archivo para {est_codigo}: {str(e)}')
            
            if archivos_subidos > 0:
                messages.success(request, f'Se subieron {archivos_subidos} archivos de examen correctamente.')
            
            # Redirigir para evitar reenvío del formulario
            return redirect('generar_reporte_secretaria', curso_codigo=curso_codigo, unidad=unidad)
        
        # GET: Generar reporte
        reporte = notasService.generarReporteSecretaria(
            curso_codigo=curso_codigo,
            unidad=unidad,
            profesor_usuario_id=profesor.usuario.codigo
        )
        
        # Obtener todas las notas parciales con información de archivos
        notas_parciales_todas = Nota.objects.filter(
            curso=curso,
            unidad=unidad,
            categoria='PARCIAL'
        ).select_related('estudiante__usuario').order_by('-valor')
        
        # Identificar las 3 notas importantes: mayor, menor y promedio
        if notas_parciales_todas.exists():
            nota_maxima = notas_parciales_todas.first()  # Mayor nota
            nota_minima = notas_parciales_todas.last()   # Menor nota
            
            # Calcular nota más cercana al promedio
            promedio = reporte['estadisticas_parcial']['promedio']
            nota_promedio = min(
                notas_parciales_todas,
                key=lambda n: abs(float(n.valor) - float(promedio))
            )
            
            # IDs de las notas importantes
            notas_importantes_ids = {nota_maxima.id, nota_minima.id, nota_promedio.id}
            
            notas_importantes = [
                {'nota': nota_maxima, 'tipo': 'MAYOR', 'descripcion': 'Nota Máxima'},
                {'nota': nota_promedio, 'tipo': 'PROMEDIO', 'descripcion': 'Nota Promedio'},
                {'nota': nota_minima, 'tipo': 'MENOR', 'descripcion': 'Nota Mínima'},
            ]
        else:
            notas_importantes = []
            notas_importantes_ids = set()
        
        context = {
            'profesor': profesor,
            'curso': curso,
            'unidad': unidad,
            'reporte': reporte,
            'notas_parciales': notas_parciales_todas,
            'notas_importantes': notas_importantes,
            'stats_parcial': reporte['estadisticas_parcial'],
            'stats_continua': reporte['estadisticas_continua'],
            'fecha_generacion': timezone.now()
        }
        
        return render(request, 'profesor/reporte_notas_secretaria.html', context)
        
    except Profesor.DoesNotExist:
        messages.error(request, 'Solo los profesores pueden acceder a esta página.')
        return redirect('login')
    except Exception as e:
        messages.error(request, f'Error: {str(e)}')
        return redirect('profesor_dashboard')


@login_required
def enviar_reporte_secretaria(request, curso_codigo, unidad):
    """
    Envía el reporte a secretaría guardándolo en la base de datos
    """
    if request.method != 'POST':
        return redirect('generar_reporte_secretaria', curso_codigo=curso_codigo, unidad=unidad)
    
    try:
        from app.models.evaluacion.models import ReporteNotas
        from app.models.usuario.models import Usuario
        
        profesor = Profesor.objects.get(usuario=request.user)
        curso = get_object_or_404(Curso, codigo=curso_codigo)
        
        # Obtener observaciones
        observaciones = request.POST.get('observaciones', '')
        
        # Generar estadísticas
        reporte_data = notasService.generarReporteSecretaria(
            curso_codigo=curso_codigo,
            unidad=unidad,
            profesor_usuario_id=profesor.usuario.codigo
        )
        
        stats_parcial = reporte_data['estadisticas_parcial']
        stats_continua = reporte_data['estadisticas_continua']
        
        # Obtener las 3 notas importantes
        notas_parciales = Nota.objects.filter(
            curso=curso,
            unidad=unidad,
            categoria='PARCIAL'
        ).select_related('estudiante').order_by('-valor')
        
        if not notas_parciales.exists():
            messages.error(request, 'No hay notas registradas para generar el reporte.')
            return redirect('generar_reporte_secretaria', curso_codigo=curso_codigo, unidad=unidad)
        
        nota_maxima = notas_parciales.first()
        nota_minima = notas_parciales.last()
        promedio = float(stats_parcial['promedio'])
        nota_promedio = min(notas_parciales, key=lambda n: abs(float(n.valor) - promedio))
        
        # Crear el reporte
        reporte = ReporteNotas.objects.create(
            curso=curso,
            unidad=unidad,
            profesor=profesor,
            observaciones=observaciones,
            # Estadísticas parcial
            promedio_parcial=stats_parcial['promedio'],
            nota_maxima_parcial=stats_parcial['nota_maxima'],
            nota_minima_parcial=stats_parcial['nota_minima'],
            aprobados_parcial=stats_parcial['aprobados'],
            desaprobados_parcial=stats_parcial['desaprobados'],
            # Estadísticas continua
            promedio_continua=stats_continua['promedio'],
            nota_maxima_continua=stats_continua['nota_maxima'],
            nota_minima_continua=stats_continua['nota_minima'],
            aprobados_continua=stats_continua['aprobados'],
            desaprobados_continua=stats_continua['desaprobados'],
            # Estudiantes
            estudiante_nota_mayor=nota_maxima.estudiante,
            estudiante_nota_menor=nota_minima.estudiante,
            estudiante_nota_promedio=nota_promedio.estudiante,
            # Archivos de examen (pueden ser None)
            examen_nota_mayor=nota_maxima.archivo_examen if nota_maxima.archivo_examen else None,
            examen_nota_menor=nota_minima.archivo_examen if nota_minima.archivo_examen else None,
            examen_nota_promedio=nota_promedio.archivo_examen if nota_promedio.archivo_examen else None,
        )
        
        messages.success(request, f'Reporte enviado exitosamente a secretaría. Código de reporte: #{reporte.id}')
        return redirect('seleccionar_curso_notas')
        
    except Profesor.DoesNotExist:
        messages.error(request, 'Solo los profesores pueden acceder a esta página.')
        return redirect('login')
    except Exception as e:
        messages.error(request, f'Error al enviar reporte: {str(e)}')
        return redirect('generar_reporte_secretaria', curso_codigo=curso_codigo, unidad=unidad)


@login_required
def ver_reporte_secretaria(request, reporte_id):
    """
    Vista para que secretaría vea el detalle de un reporte
    """
    try:
        from app.models.evaluacion.models import ReporteNotas
        
        reporte = get_object_or_404(
            ReporteNotas.objects.select_related(
                'curso',
                'profesor__usuario',
                'estudiante_nota_mayor__usuario',
                'estudiante_nota_menor__usuario',
                'estudiante_nota_promedio__usuario'
            ),
            id=reporte_id
        )
        
        context = {
            'reporte': reporte,
        }
        
        return render(request, 'secretaria/ver_reporte_detalle.html', context)
        
    except Exception as e:
        messages.error(request, f'Error al cargar el reporte: {str(e)}')
        return redirect('secretaria_reportes')


@login_required
def descargar_reporte_pdf(request, curso_codigo, unidad):
    """
    Genera y descarga reporte en PDF
    """
    try:
        # Los imports de weasyprint y render_to_string deben estar aquí si no se usan globalmente
        from django.template.loader import render_to_string
        from weasyprint import HTML
        import tempfile
        
        profesor = Profesor.objects.get(usuario=request.user)
        
        reporte = notasService.generarReporteSecretaria(
            curso_codigo=curso_codigo,
            unidad=unidad,
            profesor_usuario_id=profesor.usuario.codigo
        )
        
        # Renderizar HTML
        html_string = render_to_string('profesor/reporte_notas_pdf.html', {'reporte': reporte})
        
        # Convertir a PDF
        html = HTML(string=html_string)
        pdf = html.write_pdf()
        
        # Retornar PDF
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="reporte_notas_{curso_codigo}_U{unidad}.pdf"'
        
        return response
        
    except Exception as e:
        messages.error(request, f'Error al generar PDF: {str(e)}')
        return redirect('profesor_dashboard')


# --- Vistas para Gestión de Exámenes (NUEVAS FUNCIONALIDADES RF 5.3 y RF 5.5) ---

@login_required
def listar_fechas_examen(request, curso_codigo):
    """
    Vista GET para mostrar las fechas de exámenes programadas (RF 5.3) 
    y determinar si el usuario es Profesor Titular para habilitar la edición (RF 5.5).
    """
    from app.models.horario.models import Horario
    
    try:
        profesor = Profesor.objects.get(usuario=request.user)
        curso = get_object_or_404(Curso, codigo=curso_codigo)
        
        # 1. Determinar si es el Profesor Titular (RF 5.5)
        # Verificar si tiene horario de TEORIA en este curso
        es_titular = Horario.objects.filter(
            curso=curso,
            profesor=profesor,
            tipo_clase='TEORIA',
            is_active=True
        ).exists()
        
        # Si no es titular, verificar si puede programar exámenes
        puede_programar = es_titular
        if not es_titular:
            # Verificar si hay profesor de teoría
            hay_profesor_teoria = Horario.objects.filter(
                curso=curso,
                tipo_clase='TEORIA',
                is_active=True
            ).exists()
            
            # Si no hay profesor de teoría, verificar que sea profesor de práctica
            if not hay_profesor_teoria:
                es_profesor_practica = Horario.objects.filter(
                    curso=curso,
                    profesor=profesor,
                    tipo_clase='PRACTICA',
                    is_active=True
                ).exists()
                puede_programar = es_profesor_practica

        # 2. Obtener fechas programadas
        fechas_programadas = FechaExamen.objects.filter(
            curso=curso,
            is_active=True
        ).select_related('profesor_responsable').order_by('fecha_inicio')
        
        # 3. Obtener los choices para el formulario
        tipos_examen_choices = [
            ('PRIMER_PARCIAL', 'Primer Parcial'),
            ('SEGUNDO_PARCIAL', 'Segundo Parcial'),
            ('TERCER_PARCIAL', 'Tercer Parcial'),
        ]

        context = {
            'curso': curso,
            'es_titular': es_titular,
            'puede_programar': puede_programar,
            'fechas': fechas_programadas,
            'tipos_examen_choices': tipos_examen_choices
        }
        return render(request, 'profesor/fechas_examenes.html', context)

    except Profesor.DoesNotExist:
        messages.error(request, 'Usuario no reconocido como profesor.')
        return redirect('profesor_dashboard')
    except Exception as e:
        messages.error(request, f'Error al cargar fechas: {str(e)}')
        return redirect('profesor_dashboard')


@login_required
def programar_examen_post(request, curso_codigo):
    """
    Procesa el POST para registrar una FechaExamen (RF 5.3).
    Valida el rol de Profesor Titular (RF 5.5).
    """
    from app.models.horario.models import Horario
    
    redirect_url = 'profesor_fechas_examen'

    if request.method != 'POST':
        messages.error(request, "Solicitud inválida.")
        return redirect(redirect_url, curso_codigo=curso_codigo)

    try:
        profesor = Profesor.objects.get(usuario=request.user)
        curso = get_object_or_404(Curso, codigo=curso_codigo)

        # 1. VALIDACIÓN DE ROL CRÍTICA (RF 5.5)
        # Verificar si es titular (tiene horario de TEORIA)
        es_titular = Horario.objects.filter(
            curso=curso,
            profesor=profesor,
            tipo_clase='TEORIA',
            is_active=True
        ).exists()
        
        if not es_titular:
            # Lógica de fallback para permisos...
            hay_profesor_teoria = Horario.objects.filter(
                curso=curso,
                tipo_clase='TEORIA',
                is_active=True
            ).exists()
            
            if hay_profesor_teoria:
                messages.error(request, "Permiso denegado: Solo el profesor titular de este curso puede programar exámenes.")
                return redirect(redirect_url, curso_codigo=curso_codigo)
            else:
                es_profesor_practica = Horario.objects.filter(
                    curso=curso,
                    profesor=profesor,
                    tipo_clase='PRACTICA',
                    is_active=True
                ).exists()
                
                if not es_profesor_practica:
                    messages.error(request, "Permiso denegado: No está asignado como profesor de este curso.")
                    return redirect(redirect_url, curso_codigo=curso_codigo)
        
        # 2. Obtener y validar datos (LÓGICA ACTUALIZADA)
        tipo_examen = request.POST.get('tipo_examen')
        fecha_examen_str = request.POST.get('fecha_examen') # Solo una fecha
        observaciones = request.POST.get('observaciones', '')
        
        fecha_examen = parse_date(fecha_examen_str)
        
        if not all([fecha_examen, tipo_examen]):
            messages.error(request, "Faltan datos obligatorios para programar el examen.")
            return redirect(redirect_url, curso_codigo=curso_codigo)

        # Validación A: Fecha no pasada
        if fecha_examen < date.today():
            messages.error(request, "Error: La fecha del examen no puede ser en el pasado.")
            return redirect(redirect_url, curso_codigo=curso_codigo)

        # Validación B: Profesor tiene clase ese día
        dia_semana_examen = fecha_examen.isoweekday() # 1=Lunes, 7=Domingo
        tiene_clase_ese_dia = Horario.objects.filter(
            curso=curso,
            profesor=profesor,
            dia_semana=dia_semana_examen,
            is_active=True
        ).exists()

        if not tiene_clase_ese_dia:
            messages.error(request, f"Error: No puede programar el examen el {fecha_examen.strftime('%d/%m/%Y')} porque usted no tiene clases asignadas ese día.")
            return redirect(redirect_url, curso_codigo=curso_codigo)

        # 3. Guardar el registro de FechaExamen
        periodo_academico = getattr(curso, 'periodo_academico', f"{date.today().year}-2") 

        # Guardamos inicio y fin como el mismo día
        nueva_fecha = FechaExamen(
            curso=curso,
            tipo_examen=tipo_examen, 
            fecha_inicio=fecha_examen,
            fecha_fin=fecha_examen,
            periodo_academico=periodo_academico, 
            observaciones=observaciones,
            profesor_responsable=profesor
        )
        
        nueva_fecha.full_clean()
        nueva_fecha.save()
        
        messages.success(request, f"¡Examen {nueva_fecha.get_tipo_examen_display()} programado exitosamente para el {fecha_examen.strftime('%d/%m/%Y')}!")

    except Profesor.DoesNotExist:
        messages.error(request, "Usuario no reconocido como profesor.")
    except ValidationError as e:
        error_message = ', '.join(e.messages) if hasattr(e, 'messages') else str(e)
        messages.error(request, f"Error de validación: {error_message}")
    except Exception as e:
        messages.error(request, f"Error al guardar la fecha: {str(e)}")

    return redirect(redirect_url, curso_codigo=curso_codigo)


@login_required
def descargar_plantilla_notas(request, curso_codigo, unidad):
    """
    Genera y descarga una plantilla Excel para registro de notas
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from django.http import HttpResponse
        
        profesor = Profesor.objects.get(usuario=request.user)
        curso = get_object_or_404(Curso, codigo=curso_codigo)
        
        # Verificar que el profesor esté asignado al curso
        from app.models.horario.models import Horario
        horarios_profesor = Horario.objects.filter(
            profesor=profesor,
            curso=curso,
            is_active=True,
            periodo_academico='2025-B'
        )
        
        if not horarios_profesor.exists():
            messages.error(request, 'No está asignado a este curso.')
            return redirect('seleccionar_curso_notas')
        
        # Obtener estudiantes matriculados
        matriculas = notasService.obtenerEstudiantesParaNotas(
            curso_codigo=curso_codigo,
            profesor_usuario_id=profesor.usuario.codigo
        )
        
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Registro Notas"
        
        # Estilos
        header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        title_font = Font(bold=True, size=14)
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        # Título
        ws.merge_cells('A1:D1')
        cell = ws['A1']
        cell.value = f"Registro de Notas: {curso.nombre}"
        cell.font = title_font
        cell.alignment = center_alignment
        
        # Información del curso
        ws.merge_cells('A2:D2')
        cell = ws['A2']
        cell.value = f"Profesor: {profesor.usuario.nombres} {profesor.usuario.apellidos}"
        cell.alignment = center_alignment
        
        ws.merge_cells('A3:D3')
        cell = ws['A3']
        cell.value = f"Unidad: {unidad}"
        cell.alignment = center_alignment
        
        # Encabezados de columnas
        headers = ['CUI', 'Estudiante', 'Parcial', 'Continua']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
        
        # Datos de estudiantes
        row_num = 6
        for matricula in matriculas:
            estudiante = matricula.estudiante
            ws.cell(row=row_num, column=1, value=estudiante.codigo_estudiante)
            ws.cell(row=row_num, column=2, value=f"{estudiante.usuario.nombres} {estudiante.usuario.apellidos}")
            ws.cell(row=row_num, column=3, value="")  # Parcial vacío
            ws.cell(row=row_num, column=4, value="")  # Continua vacío
            row_num += 1
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        
        # Preparar respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="Modelo-Registro-Notas-{curso.codigo}-U{unidad}.xlsx"'
        
        wb.save(response)
        return response
        
    except Profesor.DoesNotExist:
        messages.error(request, 'Solo los profesores pueden acceder a esta página.')
        return redirect('login')
    except Exception as e:
        messages.error(request, f'Error al generar plantilla: {str(e)}')
        return redirect('ingresar_notas', curso_codigo=curso_codigo, unidad=unidad)


@login_required
def subir_notas_excel(request, curso_codigo, unidad):
    """
    Procesa la subida de notas desde archivo Excel
    """
    if request.method != 'POST':
        return redirect('ingresar_notas', curso_codigo=curso_codigo, unidad=unidad)
    
    try:
        import openpyxl
        
        profesor = Profesor.objects.get(usuario=request.user)
        curso = get_object_or_404(Curso, codigo=curso_codigo)
        
        # Verificar archivo
        archivo_excel = request.FILES.get('archivo_excel')
        if not archivo_excel:
            messages.error(request, 'Debe seleccionar un archivo Excel.')
            return redirect('ingresar_notas', curso_codigo=curso_codigo, unidad=unidad)
        
        # Leer archivo Excel
        wb = openpyxl.load_workbook(archivo_excel)
        ws = wb.active
        
        # Procesar datos (empezar desde la fila 6, después de los encabezados)
        notas_data = []
        errores = []
        
        for row_num in range(6, ws.max_row + 1):
            cui = ws.cell(row=row_num, column=1).value
            parcial = ws.cell(row=row_num, column=3).value
            continua = ws.cell(row=row_num, column=4).value
            
            if not cui:
                continue
            
            # Buscar estudiante por CUI
            try:
                estudiante = Estudiante.objects.get(codigo_estudiante=str(cui).strip())
                
                # Validar notas
                if parcial is not None:
                    try:
                        parcial_val = float(parcial)
                        if parcial_val < 0 or parcial_val > 20:
                            errores.append(f"CUI {cui}: Nota parcial fuera de rango (0-20)")
                            continue
                    except (ValueError, TypeError):
                        errores.append(f"CUI {cui}: Nota parcial inválida")
                        continue
                else:
                    parcial_val = None
                
                if continua is not None:
                    try:
                        continua_val = float(continua)
                        if continua_val < 0 or continua_val > 20:
                            errores.append(f"CUI {cui}: Nota continua fuera de rango (0-20)")
                            continue
                    except (ValueError, TypeError):
                        errores.append(f"CUI {cui}: Nota continua inválida")
                        continue
                else:
                    continua_val = None
                
                if parcial_val is not None or continua_val is not None:
                    notas_data.append({
                        'estudiante_codigo': estudiante.usuario.codigo,
                        'nota_parcial': parcial_val,
                        'nota_continua': continua_val,
                        'archivo_examen': None
                    })
                
            except Estudiante.DoesNotExist:
                errores.append(f"CUI {cui}: Estudiante no encontrado")
        
        # Ingresar notas usando el servicio
        if notas_data:
            resultado = notasService.ingresarNotas(
                curso_codigo=curso_codigo,
                profesor_usuario_id=profesor.usuario.codigo,
                unidad=unidad,
                notas_data=notas_data
            )
            
            if resultado['success']:
                messages.success(
                    request, 
                    f'Notas cargadas desde Excel: {resultado["notas_creadas"]} creadas, {resultado["notas_actualizadas"]} actualizadas.'
                )
                
                # Mostrar errores si los hay
                for error in errores:
                    messages.warning(request, error)
                
                # Establecer flag para mostrar pop-up de recordatorio
                request.session['mostrar_recordatorio_notas'] = True
            else:
                messages.error(request, 'Error al procesar notas del Excel.')
        else:
            messages.warning(request, 'No se encontraron notas válidas en el archivo Excel.')
            for error in errores:
                messages.warning(request, error)
        
        return redirect('ingresar_notas', curso_codigo=curso_codigo, unidad=unidad)
        
    except Profesor.DoesNotExist:
        messages.error(request, 'Solo los profesores pueden acceder a esta página.')
        return redirect('login')
    except Exception as e:
        messages.error(request, f'Error al procesar archivo Excel: {str(e)}')
        return redirect('ingresar_notas', curso_codigo=curso_codigo, unidad=unidad)